import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def setup_sheet(sheet, title, columns):
    # Header styling
    header_font = Font(bold=True, size=10)
    regular_font = Font(size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title row
    sheet['A1'] = title
    sheet.merge_cells('A1:P1')
    sheet['A1'].font = header_font
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet['A1'].border = border
    
    # Add A-End and Z-End labels for all sheets
    # A-End label (Columns A to H)
    sheet.merge_cells('A2:H2')
    sheet['A2'] = "A-End"
    sheet['A2'].fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    sheet['A2'].font = Font(color="FFFFFF", bold=True, size=10)
    sheet['A2'].alignment = Alignment(horizontal='center')
    
    # Z-End label (Columns I to O)
    sheet.merge_cells('I2:O2')
    sheet['I2'] = "Z-End"
    sheet['I2'].fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    sheet['I2'].font = Font(color="FFFFFF", bold=True, size=10)
    sheet['I2'].alignment = Alignment(horizontal='center')
    
    # Fiber specification text
    fiber_spec = "Line Fiber - Single Mode Simplex / Duplex LC/LC bend insensitive G657A (RED in Colour) (2MM Jacket)"
    sheet['A3'] = fiber_spec
    sheet.merge_cells('A3:P3')
    sheet['A3'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    sheet['A3'].font = Font(color="FFFFFF", bold=True, size=10)
    sheet['A3'].alignment = Alignment(horizontal='center')
    
    # Column headers
    for col, header in enumerate(columns, 1):
        cell = sheet.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Add empty rows for data
    for row in range(5, 25):
        for col in range(1, len(columns) + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
            cell.font = regular_font

def populate_terminal_sheet(sheet, site_name, system_name, is_terminal_a, adjacent_ila):
    # Device names
    device_name = f"{site_name}-wdm-{system_name}-s01"
    sheet.cell(row=5, column=1, value=device_name)
    sheet.cell(row=6, column=1, value=device_name)
    
    # Rack Location (Column B)
    for row in range(5, 7):
        sheet.cell(row=row, column=2, value="TBC")
    
    # Rack RU# (Column C)
    for row in range(5, 7):
        sheet.cell(row=row, column=3, value=33)
    
    # Slot (Column D)
    for row in range(5, 7):
        sheet.cell(row=row, column=4, value=2)
    
    # Port (Column E)
    sheet.cell(row=5, column=5, value="Line Out 5")
    sheet.cell(row=6, column=5, value="Line In 6")
    
    # Pluggable (Column F)
    for row in range(5, 7):
        sheet.cell(row=row, column=6, value="n/a")
    
    # Notes (Column G)
    for row in range(5, 7):
        sheet.cell(row=row, column=7, value="Simplex Fiber")
    
    # A End Fiber Label (Column H)
    if is_terminal_a:
        sheet.cell(row=5, column=8, value=f"@TBC U33 i01-2-5 : none TBC Fiber none (Tx to {adjacent_ila})")
        sheet.cell(row=6, column=8, value=f"@TBC U33 i01-2-6 : none TBC Fiber none (Rx from {adjacent_ila})")
    else:
        sheet.cell(row=5, column=8, value=f"@TBC U33 i01-2-5 : none TBC Fiber none (Tx to {adjacent_ila})")
        sheet.cell(row=6, column=8, value=f"@TBC U33 i01-2-6 : none TBC Fiber none (Rx from {adjacent_ila})")
    
    # Patch Panel (Column I)
    for row in range(5, 7):
        sheet.cell(row=row, column=9, value="TBC")
    
    # Rack Location to Pluggable (Columns J to N)
    for row in range(5, 7):
        sheet.cell(row=row, column=10, value="None")  # Rack Location
        sheet.cell(row=row, column=11, value="None")  # Rack RU#
        sheet.cell(row=row, column=12, value="None")  # Slot
        sheet.cell(row=row, column=13, value="None")  # Port/Fiber
        sheet.cell(row=row, column=14, value="N/A")   # Pluggable
    
    # Z End Fiber Label (Column O)
    sheet.cell(row=5, column=15, value="none TBC Fiber none : TBC U33 i01-2-5")
    sheet.cell(row=6, column=15, value="none TBC Fiber none : TBC U33 i01-2-6")
    
    # Notes (Column P)
    if is_terminal_a:
        sheet.cell(row=5, column=16, value=adjacent_ila)
        sheet.cell(row=6, column=16, value=adjacent_ila)
    else:
        sheet.cell(row=5, column=16, value=adjacent_ila)
        sheet.cell(row=6, column=16, value=adjacent_ila)

def populate_ila_sheet(sheet, site_name, system_name, prev_site, next_site):
    # Device names
    device_name = f"{site_name}-wdm-{system_name}-i01"
    for row in range(5, 9):
        sheet.cell(row=row, column=1, value=device_name)
    
    # Rack Location (Column B)
    for row in range(5, 9):
        sheet.cell(row=row, column=2, value="TBC")
    
    # Rack RU# (Column C)
    for row in range(5, 9):
        sheet.cell(row=row, column=3, value=33)
    
    # Slot (Column D)
    for row in range(5, 7):
        sheet.cell(row=row, column=4, value=4)
    for row in range(7, 9):
        sheet.cell(row=row, column=4, value=6)
    
    # Port (Column E)
    port_values = ["Line 1 Out 5", "Line 1 In 6", "Line 2 Out 5", "Line 2 In 6"]
    for row, value in enumerate(port_values, 5):
        sheet.cell(row=row, column=5, value=value)
    
    # Pluggable (Column F)
    for row in range(5, 9):
        sheet.cell(row=row, column=6, value="n/a")
    
    # Notes (Column G)
    for row in range(5, 9):
        sheet.cell(row=row, column=7, value="Simplex Fiber")
    
    # A End Fiber Label (Column H)
    sheet.cell(row=5, column=8, value=f"@TBC U33 i01-2-5 : none TBC Fiber none (Tx to {prev_site})")
    sheet.cell(row=6, column=8, value=f"@TBC U33 i01-2-6 : none TBC Fiber none (Rx from {prev_site})")
    sheet.cell(row=7, column=8, value=f"@TBC U33 i01-7-5 : none TBC Fiber none (Tx to {next_site})")
    sheet.cell(row=8, column=8, value=f"@TBC U33 i01-7-6 : none TBC Fiber none (Rx from {next_site})")
    
    # Patch Panel (Column I)
    for row in range(5, 9):
        sheet.cell(row=row, column=9, value="TBC")
    
    # Rack Location to Pluggable (Columns J to N)
    for row in range(5, 9):
        sheet.cell(row=row, column=10, value="None")  # Rack Location
        sheet.cell(row=row, column=11, value="None")  # Rack RU#
        sheet.cell(row=row, column=12, value="None")  # Slot
        sheet.cell(row=row, column=13, value="None")  # Port/Fiber
        sheet.cell(row=row, column=14, value="N/A")   # Pluggable
    
    # Z End Fiber Label (Column O)
    z_end_labels = [
        "none TBC Fiber none : TBC U33 i01-2-5",
        "none TBC Fiber none : TBC U33 i01-2-6",
        "none TBC Fiber none : TBC U33 i01-7-5",
        "none TBC Fiber none : TBC U33 i01-7-6"
    ]
    for row, label in enumerate(z_end_labels, 5):
        sheet.cell(row=row, column=15, value=label)
    
    # Notes (Column P)
    sheet.cell(row=5, column=16, value=prev_site)
    sheet.cell(row=6, column=16, value=prev_site)
    sheet.cell(row=7, column=16, value=next_site)
    sheet.cell(row=8, column=16, value=next_site)

def create_fiber_doc(system_name, terminal_a, terminal_b, ila_sites, output_path):
    wb = openpyxl.Workbook()
    columns = [
        'Device',
        'Rack Location',
        'Rack RU#',
        'Slot',
        'Port',
        'Pluggable',
        'Notes',
        'A End Fiber Label',
        'Patch Panel',
        'Rack Location',
        'Rack RU#',
        'Slot',
        'Port/Fiber',
        'Pluggable',
        'Z End Fiber Label',
        'Notes'
    ]
    
    # Create Terminal A sheet (uppercase for sheet name)
    term_a = wb.active
    term_a.title = f"{terminal_a.upper()} TERM-R"
    setup_sheet(term_a, f"{system_name} - {terminal_a.lower()}", columns)
    populate_terminal_sheet(term_a, terminal_a.lower(), system_name, True, ila_sites[0].lower())
    
    # Create ILA sheets with proper site relationships
    for i, ila in enumerate(ila_sites):
        ila_sheet = wb.create_sheet(f"{ila.upper()} ILA-R")
        setup_sheet(ila_sheet, f"{system_name} - {ila.lower()}", columns)
        
        # For first ILA, use Terminal A as previous site
        if i == 0:
            prev_site = terminal_a.lower()
        else:
            prev_site = ila_sites[i-1].lower()
            
        # For last ILA, use Terminal B as next site
        if i == len(ila_sites) - 1:
            next_site = terminal_b.lower()
        else:
            next_site = ila_sites[i+1].lower()
        
        populate_ila_sheet(ila_sheet, ila.lower(), system_name, prev_site, next_site)
    
    # Create Terminal B sheet (uppercase for sheet name)
    term_b = wb.create_sheet(f"{terminal_b.upper()} TERM-R")
    setup_sheet(term_b, f"{system_name} - {terminal_b.lower()}", columns)
    populate_terminal_sheet(term_b, terminal_b.lower(), system_name, False, ila_sites[-1].lower())
    
    # Adjust column widths
    for sheet in wb.sheetnames:
        # Set specific column widths
        wb[sheet].column_dimensions['A'].width = 19
        wb[sheet].column_dimensions['H'].width = 46
        wb[sheet].column_dimensions['O'].width = 30
        
        # Set other columns to width 15
        for col in range(2, len(columns) + 1):
            if col != 8 and col != 15:  # Skip H and O as they're already set
                wb[sheet].column_dimensions[get_column_letter(col)].width = 15
    
    wb.save(output_path)

if __name__ == "__main__":
    # Get System Name
    system_name = input("Enter System Name: ").strip().upper()
    
    # Get Terminal A site
    terminal_a = input("Enter Terminal A site name: ").strip().upper()
    
    # Get number of ILA sites
    while True:
        try:
            num_ilas = int(input("Enter number of ILA sites: "))
            break
        except ValueError:
            print("Please enter a valid number")
    
    # Get ILA sites as comma-separated list
    while True:
        ila_input = input(f"Enter {num_ilas} ILA site names (comma-separated): ").strip().upper()
        ila_sites = [site.strip() for site in ila_input.split(',')]
        
        if len(ila_sites) != num_ilas:
            print(f"Error: You specified {num_ilas} ILA sites but provided {len(ila_sites)} names.")
            print("Please try again.")
            continue
        
        if '' in ila_sites:
            print("Error: Empty site names are not allowed.")
            print("Please try again.")
            continue
        
        if len(set(ila_sites)) != len(ila_sites):
            print("Error: Duplicate site names are not allowed.")
            print("Please try again.")
            continue
            
        break
    
    # Get Terminal B site
    terminal_b = input("Enter Terminal B site name: ").strip().upper()
    
    # Set specific output path
    output_folder = r"C:\Users\ocallagz\Desktop\Automated_Line_Cutsheet_Output"
    output_filename = f"{system_name}_{terminal_a}_{terminal_b}_line_cutsheet.xlsx"
    output_path = os.path.join(output_folder, output_filename)
    
    # Create directory if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)
    
    try:
        create_fiber_doc(system_name, terminal_a, terminal_b, ila_sites, output_path)
        print(f"\nFile created successfully at: {output_path}")
    except PermissionError:
        print("\nError: Unable to create file. Please check if you have permission to write to this location or if the file is already open.")
    except Exception as e:
        print(f"\nAn error occurred: {str(e)}")
